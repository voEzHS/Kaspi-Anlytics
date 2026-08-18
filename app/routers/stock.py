"""Остатки CRM (torgstore.zymyran.com, "Экспорт товаров с надстройками") —
загрузка снимка склада, используется вкладкой «Закуп».

Формат подтверждён вживую 05.08.2026 (см. CRM_Logistika_Sklady.md и Замечание
7 в истории Plan_Zakupa): колонки — SKU, Название, Статус, Цена, 4 продаваемых
на Kaspi склада (Первомай/Астана/Шымкент/Туздыбастау), плюс произвольный набор
"пайплайн"-столбцов (YMC-хабы, 2_ordered, 3_left_factory — набор и порядок
может меняться от выгрузки к выгрузке), плюс MAS/Общий объём — это НЕ сток
(товарная характеристика, объём/литраж), не читаем их вообще.
"""
import os
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.models import StockRow, StockUpload
from app.routers.uploads import require_admin, require_admin_or_ingest, MAX_UPLOAD_BYTES

router = APIRouter(prefix="/api/v1/stock", tags=["stock"])

# Явно опознаваемые колонки. Всё, что не сюда — считается пайплайном
# (см. ymc_transit) и суммируется, КРОМЕ явно исключённых ниже.
_COL_SKU = "sku"
_COL_NAME = "название"
_COL_STATUS = "статус"
_COL_PRICE = "цена"
_COL_WH = {
    "первомай": "wh_pervomay",
    "астана": "wh_astana",
    "шымкент": "wh_shymkent",
    "туздыбастау": "wh_tuzdybastau",
}
# "ordered" встречается в латинице внутри заголовка "2_ordered (已订购, заказаны)"
_ORDERED_HINT = "ordered"
# Явно НЕ сток — товарные характеристики/непонятные пустые поля, не суммируем
# никуда (проверено вживую: "Объём" — характеристика товара (литраж), не остаток;
# "MAS" — всегда пусто в выгрузке).
_EXCLUDE_HINTS = ("объем", "объём", "mas")


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def parse_stock_excel(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return []

    # Заголовок почти всегда строка 1, но на всякий случай ищем строку с "sku"
    header_idx = 0
    for i, row in enumerate(rows_raw[:5]):
        if any(_norm(c) == "sku" for c in row):
            header_idx = i
            break
    headers = rows_raw[header_idx]

    col_sku = col_name = col_status = col_price = col_ordered = None
    col_wh: dict[str, int] = {}
    transit_cols: list[int] = []

    for idx, h in enumerate(headers):
        n = _norm(h)
        if not n:
            continue
        if n == _COL_SKU:
            col_sku = idx
        elif _COL_NAME in n:
            col_name = idx
        elif _COL_STATUS in n:
            col_status = idx
        elif _COL_PRICE in n:
            col_price = idx
        elif any(k in n for k in _COL_WH):
            for k, field in _COL_WH.items():
                if k in n:
                    col_wh[field] = idx
                    break
        elif _ORDERED_HINT in n:
            col_ordered = idx
        elif any(k in n for k in _EXCLUDE_HINTS):
            continue  # товарная характеристика — не сток, никуда не суммируем
        else:
            transit_cols.append(idx)

    if col_sku is None:
        raise ValueError("Колонка SKU не найдена в файле")

    result = []
    for row in rows_raw[header_idx + 1:]:
        if not row or col_sku >= len(row):
            continue
        sku_raw = row[col_sku]
        if sku_raw is None or str(sku_raw).strip() == "":
            continue
        sku = str(sku_raw).strip().upper()

        def get(i):
            return row[i] if i is not None and i < len(row) else None

        transit_sum = sum(_num(get(i)) for i in transit_cols)

        result.append({
            "sku": sku,
            "name": str(get(col_name) or "").strip(),
            "status": str(get(col_status) or "").strip(),
            "price": _num(get(col_price)),
            "wh_pervomay": _num(get(col_wh.get("wh_pervomay"))),
            "wh_astana": _num(get(col_wh.get("wh_astana"))),
            "wh_shymkent": _num(get(col_wh.get("wh_shymkent"))),
            "wh_tuzdybastau": _num(get(col_wh.get("wh_tuzdybastau"))),
            "ymc_transit": transit_sum,
            "ordered": _num(get(col_ordered)),
        })
    return result


@router.post("/", summary="Загрузить снимок остатков (заменяет предыдущий целиком)")
async def upload_stock(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_admin_or_ingest),
):
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)")

    import tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        parsed = parse_stock_excel(tmp_path)
    except Exception as e:
        raise HTTPException(422, f"Ошибка разбора файла: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not parsed:
        raise HTTPException(422, "В файле не найдено строк с данными")

    # Остатки — не накопительные (это состояние "на сейчас"), поэтому старый
    # снимок удаляется целиком перед вставкой нового, а не суммируется с ним.
    old_uploads = (await db.execute(select(StockUpload))).scalars().all()
    for u in old_uploads:
        await db.delete(u)
    await db.flush()

    upload = StockUpload(filename=file.filename, row_count=len(parsed))
    db.add(upload)
    await db.flush()

    CHUNK = 500
    for i in range(0, len(parsed), CHUNK):
        chunk = parsed[i: i + CHUNK]
        db.add_all([StockRow(upload_id=upload.id, **row) for row in chunk])

    await db.commit()
    await db.refresh(upload)

    return {
        "id": upload.id,
        "filename": upload.filename,
        "row_count": upload.row_count,
        "created_at": upload.created_at.isoformat(),
        "replaced_previous": len(old_uploads) > 0,
    }


@router.get("/", summary="Текущий снимок остатков (метаданные)")
async def get_stock_status(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(StockUpload).order_by(StockUpload.created_at.desc()))
    upload = result.scalars().first()
    if not upload:
        return {"loaded": False}
    return {
        "loaded": True,
        "id": upload.id,
        "filename": upload.filename,
        "row_count": upload.row_count,
        "created_at": upload.created_at.isoformat(),
    }


@router.delete("/", summary="Удалить текущий снимок остатков")
async def delete_stock(db: AsyncSession = Depends(get_db), _: None = Depends(require_admin)):
    old_uploads = (await db.execute(select(StockUpload))).scalars().all()
    for u in old_uploads:
        await db.delete(u)
    await db.commit()
    return {"deleted": len(old_uploads)}
