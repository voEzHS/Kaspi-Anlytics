"""Excel upload & parsing router."""
import glob
import os
import re
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Request, UploadFile
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.models import DeptEnum, KaspiRow, Upload
from app.analytics.tip_classifier import classify_rows

router = APIRouter(prefix="/api/v1/uploads", tags=["uploads"])

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
# Must match main.py's default — used only to recognize the admin account
# that already passed the front-door Basic Auth check (see require_admin).
BASIC_AUTH_USER = os.getenv("BASIC_AUTH_USER", "torgstore")


async def require_admin(request: Request, x_admin_token: Optional[str] = Header(None)):
    """Allow write operations only if the caller is the admin.

    Two independent ways to prove that:
      1. The in-app x-admin-token header matches ADMIN_PASSWORD (the
         password modal in the UI).
      2. The request already passed the front-door Basic Auth middleware
         AS the admin account (BASIC_AUTH_USER). basic_auth_middleware
         stashes this on request.state before the route ever runs, so a
         logged-in admin never has to re-type the same password into a
         second, separate prompt just to upload or delete data.
    """
    if not ADMIN_PASSWORD:
        return  # No password set — open access (dev mode)
    if getattr(request.state, "basic_auth_user", None) == BASIC_AUTH_USER:
        return
    if x_admin_token != ADMIN_PASSWORD:
        raise HTTPException(403, "Неверный пароль администратора")

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/app/uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_MB", "50")) * 1024 * 1024  # default 50 MB

# ── Column name → field key mapping ──────────────────────────────────────────
HEADER_MAP = {
    # Код / Артикул
    "код товара": "kod", "код": "kod", "код тов.": "kod",
    "артикул": "kod", "артикул товара": "kod",
    "id товара": "kod", "sku": "kod",
    # Тип
    "тип": "tip", "тип товара": "tip",
    # Название / Наименование
    "название товара": "name", "название": "name",
    "наименование": "name", "наименование товара": "name",
    "товар": "name",
    # Бренд
    "бренд": "brand", "марка": "brand", "производитель": "brand",
    # Объём
    "объем": "volume", "объём": "volume", "обем": "volume",
    # Ветка
    "ветка": "vetka", "категория": "vetka", "подкатегория": "vetka",
    # Цвет
    "цвет": "color",
    # Месяц
    "месяц": "month", "период": "month",
    # РРЦ
    "ррц": "rrc", "цена": "rrc", "рекомендованная цена": "rrc",
    # Штуки — разные варианты
    "штуки": "units", "шт.": "units", "шт": "units",
    "продажи шт": "units", "кол-во": "units", "количество": "units",
    "в штуках": "units", "штуки ": "units",
    # Выручка
    "выручка": "revenue", "оборот": "revenue", "продажи": "revenue",
    # ABC
    "abc": "abc", "abc анализ": "abc",
    # Продавцов
    "продавцов": "sellers", "продавцы": "sellers",
    # Рейтинг
    "рейтинг товара": "rating", "рейтинг": "rating", "оценка": "rating",
    # Отзывы
    "отзывы": "reviews", "отзывов": "reviews",
    # Размороз
    "размороз": "thaw",
}

# Нормализация названий месяцев к полной форме (с заглавной)
_MONTH_NORM: dict[str, str] = {
    "январь": "Январь", "янв": "Январь", "january": "Январь", "jan": "Январь",
    "февраль": "Февраль", "фев": "Февраль", "february": "Февраль", "feb": "Февраль",
    "март": "Март", "mar": "Март", "march": "Март",
    "апрель": "Апрель", "апр": "Апрель", "april": "Апрель", "apr": "Апрель",
    "май": "Май", "may": "Май",
    "июнь": "Июнь", "июн": "Июнь", "june": "Июнь", "jun": "Июнь",
    "июль": "Июль", "июл": "Июль", "july": "Июль", "jul": "Июль",
    "август": "Август", "авг": "Август", "august": "Август", "aug": "Август",
    "сентябрь": "Сентябрь", "сен": "Сентябрь", "september": "Сентябрь", "sep": "Сентябрь",
    "октябрь": "Октябрь", "окт": "Октябрь", "october": "Октябрь", "oct": "Октябрь",
    "ноябрь": "Ноябрь", "ноя": "Ноябрь", "november": "Ноябрь", "nov": "Ноябрь",
    "декабрь": "Декабрь", "дек": "Декабрь", "december": "Декабрь", "dec": "Декабрь",
}


def _norm_month(raw: str) -> str:
    """Normalize month string to full Russian name with capital letter."""
    if not raw:
        return raw
    key = raw.strip().lower()
    return _MONTH_NORM.get(key, raw.strip().capitalize())


def _norm(s: str) -> str:
    return re.sub(r"[\s,]+", " ", str(s or "").strip().lower())


def _pn(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def _find_sheet(wb) -> str:
    """Find the main data sheet by name (case-insensitive prefix match)."""
    names = wb.sheetnames
    # Priority: sheet whose name starts with "главн" (Главная, Главный, Главн...)
    for s in names:
        if s.strip().lower().startswith("главн"):
            return s
    # Fallback: sheet with most rows (pick the first non-empty one)
    return names[0]


def _find_header(rows_raw: list) -> int:
    """
    Find the row index of the column header by scanning up to 60 rows.

    Strategy: require BOTH a product-identifier keyword (код / название / наименование)
    AND a brand/metric keyword in the same row.  This prevents matching summary/pivot
    tables at the top of the sheet that have 'Бренд' / 'Выручка' but no product ids.
    """
    id_kw    = {"код", "название", "наименование", "артикул", "name", "sku"}
    brand_kw = {"бренд", "brand", "марка"}

    # Pass 1 — strict: both id + brand present
    for i, row in enumerate(rows_raw[:60]):
        joined = "|".join(str(c or "").lower() for c in row)
        if any(k in joined for k in id_kw) and any(k in joined for k in brand_kw):
            return i

    # Pass 2 — relaxed: any metric-like keyword (handles unusual formats)
    metric_kw = {"ветка", "выручка", "revenue", "штук"}
    for i, row in enumerate(rows_raw[:60]):
        joined = "|".join(str(c or "").lower() for c in row)
        if any(k in joined for k in brand_kw) and any(k in joined for k in metric_kw):
            return i

    return min(16, len(rows_raw) - 1)  # fallback


def _build_col_map(headers: tuple) -> dict[str, int]:
    """Map field names to column indices from the header row."""
    col: dict[str, int] = {}
    for idx, h in enumerate(headers):
        n = _norm(h)
        # Exact matches first — and if one fires, this column's identity is
        # fully resolved, so skip the fuzzy loop below entirely for it.
        #
        # БАГ (найден в UX-аудите 29.07, подтверждено на реальных файлах):
        # раньше фаззи-цикл ВСЕГДА выполнялся, даже после точного совпадения.
        # Для заголовка "КОД товара" точное совпадение верно ставило col["kod"],
        # но затем тот же текст "код товара" содержит подстроку "товар", и
        # фаззи-правило ("товар" → "name") перехватывало ЭТУ ЖЕ колонку под
        # "name" (т.к. поле "name" ещё не было занято на этой итерации). Когда
        # цикл доходил до настоящего "Название товара", col["name"] уже был
        # занят индексом колонки "Код", и реальное имя товара никогда не
        # читалось — get("name") и get("kod") указывали на одну ячейку.
        # Итог: "Название" в интерфейсе показывало код товара на каждой строке.
        if n in HEADER_MAP:
            field = HEADER_MAP[n]
            if field not in col:
                col[field] = idx
            continue
        # Fuzzy fallbacks — ordered from most specific to least
        for keyword, field in [
            ("бренд",    "brand"),
            ("марк",     "brand"),
            ("ветк",     "vetka"),
            ("катег",    "vetka"),
            ("штук",     "units"),
            ("в штуках", "units"),
            ("кол-во",   "units"),
            ("количест", "units"),
            ("выручк",   "revenue"),
            ("оборот",   "revenue"),
            ("месяц",    "month"),
            ("период",   "month"),
            ("ррц",      "rrc"),
            ("рейтинг",  "rating"),
            ("оценк",    "rating"),
            ("продавц",  "sellers"),
            ("отзыв",    "reviews"),
            ("назван",   "name"),   # название
            ("наимен",   "name"),   # наименование  ← КЛЮЧЕВОЙ ФИX
            ("товар",    "name"),
            ("объем",    "volume"),
            ("объём",    "volume"),
            ("abc",      "abc"),
            ("артикул",  "kod"),    # артикул → kod ← КЛЮЧЕВОЙ ФИX
            ("код товар", "kod"),
            ("id товар",  "kod"),
        ]:
            if keyword in n and field not in col:
                col[field] = idx
        # "код" только как отдельное слово (избегаем штрих-код, QR-код)
        if "kod" not in col:
            if n == "код" or n.startswith("код ") or n.endswith(" код"):
                col["kod"] = idx
    return col


# Слова-заголовки — если встречаем строку где "бренд" стоит вместо названия бренда,
# это повторный заголовок новой секции, пропускаем его
_HEADER_WORDS = {"бренд", "brand", "марка", "бренды", "наименование"}


def parse_excel(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = _find_sheet(wb)
    ws = wb[sheet_name]

    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    # Находим ПЕРВЫЙ заголовок — он задаёт маппинг колонок
    header_idx = _find_header(rows_raw)
    headers = rows_raw[header_idx]
    col = _build_col_map(headers)

    brand_col = col.get("brand")

    result = []
    # Читаем ВСЕ строки от первого заголовка до конца файла
    for row in rows_raw[header_idx + 1:]:
        # Пустая строка — пропуск
        if not row or all(c is None or c == "" for c in row):
            continue

        # Получаем значение в колонке бренда
        if brand_col is not None and brand_col < len(row):
            brand_raw = str(row[brand_col] or "").strip()
        else:
            brand_raw = ""

        # Повторный заголовок секции (бренд написан "бренд" вместо названия) — пропускаем
        if brand_raw.lower() in _HEADER_WORDS:
            continue

        brand = brand_raw.upper()
        if not brand:
            continue

        def get(field: str, _row=row):
            i = col.get(field)
            return _row[i] if i is not None and i < len(_row) else None

        revenue = _pn(get("revenue"))
        units = _pn(get("units"))
        name = str(get("name") or "").strip()
        kod = str(get("kod") or "").strip()

        # Пропускаем только полностью пустые строки (нет продаж И нет идентификатора)
        if revenue == 0 and units == 0 and not name and not kod:
            continue

        result.append({
            "kod": kod,
            "tip": str(get("tip") or "").strip(),
            "name": name,
            "brand": brand,
            "volume": str(get("volume") or "").strip(),
            "vetka": str(get("vetka") or "").strip(),
            "color": str(get("color") or "").strip(),
            "month": _norm_month(str(get("month") or "").strip()),
            "rrc": _pn(get("rrc")),
            "units": units,
            "revenue": revenue,
            "abc": str(get("abc") or "").strip().upper(),
            "sellers": _pn(get("sellers")),
            "rating": _pn(get("rating")),
            "reviews": _pn(get("reviews")),
            "thaw": _pn(get("thaw")),
        })

    return result


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/preview", summary="Preview Excel parse without saving")
async def preview_parse(file: UploadFile = File(...)):
    """Parse Excel and return diagnostic info — does NOT save to DB."""
    import tempfile
    import openpyxl as ox

    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        wb = ox.load_workbook(tmp_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        chosen_sheet = _find_sheet(wb)
        ws = wb[chosen_sheet]
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()

        header_idx = _find_header(rows_raw)
        headers = rows_raw[header_idx]
        col = _build_col_map(headers)

        sample = []
        for row in rows_raw[header_idx + 1: header_idx + 6]:
            if row and not all(c is None or c == "" for c in row):
                sample.append(list(row))

        parsed = parse_excel(tmp_path)
        brands_found = {}
        for r in parsed:
            b = r["brand"]
            brands_found[b] = brands_found.get(b, 0) + 1

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return {
        "sheets_in_file": sheets,
        "sheet_chosen": chosen_sheet,
        "total_rows_in_sheet": len(rows_raw),
        "header_row_index": header_idx,
        "header_row_content": list(headers),
        "columns_detected": col,
        "rows_parsed_total": len(parsed),
        "brands_found": dict(sorted(brands_found.items(), key=lambda x: -x[1])[:30]),
        "sample_data_rows": sample,
    }


@router.get("/auth-check", summary="Verify admin password")
async def auth_check(request: Request, x_admin_token: Optional[str] = Header(None)):
    """Returns 200 if password is correct (or already admin via Basic Auth), 403 if not."""
    if not ADMIN_PASSWORD:
        return {"ok": True}
    if getattr(request.state, "basic_auth_user", None) == BASIC_AUTH_USER:
        return {"ok": True}
    if x_admin_token != ADMIN_PASSWORD:
        raise HTTPException(403, "Неверный пароль")
    return {"ok": True}


@router.post("/", summary="Upload Excel matrix")
async def upload_file(
    file: UploadFile = File(...),
    department: str = Form(...),   # "freezers" | "refrigerated" | "ovens" | "ice_makers"
    month: Optional[str] = Form(None),   # override month for all rows
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_admin),
):
    if department not in DeptEnum.__members__:
        raise HTTPException(400, f"Unknown department: {department}")

    # Save file
    safe_name = f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)")
    with open(filepath, "wb") as f:
        f.write(content)

    # Parse
    try:
        parsed = parse_excel(filepath)
    except Exception as e:
        os.remove(filepath)
        raise HTTPException(422, f"Parse error: {e}")

    if not parsed:
        os.remove(filepath)
        raise HTTPException(422, "No data rows found. Check file format.")

    # Override month if provided
    month_clean = month.strip() if month and month.strip() else None
    if month_clean:
        for row in parsed:
            row["month"] = month_clean

    dept = DeptEnum[department]

    # ── Auto-classify missing Тип against this department's own history ────
    # Some sources (e.g. algatop for refrigerated) stopped sending a Тип
    # column entirely. Never guess: resolve only rows with hard evidence
    # (exact kod match, or a validated name/brand pattern — see
    # tip_classifier.py), and leave the rest blank for manual review via
    # GET/PATCH /api/v1/uploads/unclassified.
    tip_classification = None
    if any(not r["tip"] for r in parsed):
        hist_result = await db.execute(
            select(KaspiRow.kod, KaspiRow.tip, KaspiRow.name, KaspiRow.brand)
            .where(KaspiRow.department == dept, KaspiRow.tip.isnot(None), KaspiRow.tip != "")
        )
        historical = [tuple(r) for r in hist_result.all()]
        if historical:
            summary = classify_rows(parsed, historical)
            tip_classification = {
                "resolved_exact_kod": summary["exact_kod"],
                "resolved_exact_name": summary["exact_name_nocolor"],
                "resolved_brand_pattern": summary["brand_prefix"],
                "unresolved_count": len(summary["unresolved"]),
                "unresolved_sample": [
                    {"kod": r["kod"], "name": r["name"], "brand": r["brand"], "revenue": r["revenue"]}
                    for r in summary["unresolved"][:20]
                ],
            }

    months = sorted({r["month"] for r in parsed if r["month"]})
    subtypes = sorted({r["tip"] for r in parsed if r["tip"]})

    # ── Duplicate month guard ────────────────────────────────────────────
    # Check if any of the months in this file are already loaded for this dept.
    # If so, warn — uploading would double all metrics for those months.
    if months:
        existing_uploads = await db.execute(
            select(Upload).where(Upload.department == dept)
        )
        all_uploads = existing_uploads.scalars().all()
        already_loaded = []
        for u in all_uploads:
            for m in (u.months or []):
                if m in months:
                    already_loaded.append(m)
        if already_loaded:
            dupes = sorted(set(already_loaded))
            raise HTTPException(
                409,
                f"Данные за {', '.join(dupes)} уже загружены для этого отдела. "
                f"Удалите старую загрузку прежде чем загружать снова, иначе метрики удвоятся."
            )

    # Create upload record
    upload = Upload(
        filename=file.filename,
        department=dept,
        row_count=len(parsed),
        months=months,
        subtypes=subtypes,
    )
    db.add(upload)
    await db.flush()

    # Bulk insert rows (chunked)
    CHUNK = 500
    for i in range(0, len(parsed), CHUNK):
        chunk = parsed[i: i + CHUNK]
        db.add_all([
            KaspiRow(upload_id=upload.id, department=dept, **row)
            for row in chunk
        ])

    await db.commit()
    await db.refresh(upload)

    return {
        "id": upload.id,
        "filename": upload.filename,
        "department": department,
        "row_count": upload.row_count,
        "months": upload.months,
        "subtypes": upload.subtypes,
        "created_at": upload.created_at.isoformat(),
        "tip_classification": tip_classification,
    }


@router.post("/backfill-names", summary="Fix name==kod rows caused by the old column-mapping bug")
async def backfill_names(
    file: UploadFile = File(...),
    department: str = Form(...),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_admin),
):
    """
    UX-аудит 29.07 нашёл, что "Название" на Товарах/ABC/Приоритетах показывало
    код товара вместо имени. Причина — баг в _build_col_map (см. комментарий
    там): для строк, загруженных ДО фикса, поле name в БД уже сохранено как
    копия kod. Этот endpoint НЕ трогает данные заново (не создаёт новый Upload,
    не пересчитывает метрики) — он только допроверяет по kod уже загруженный
    файл этого отдела и обновляет name у существующих строк, где он сейчас
    пуст ИЛИ совпадает с kod (сигнатура именно этого бага). Строки, где name
    уже отличается от kod, не трогаются — чтобы не затереть ничего, что могло
    быть верным.
    """
    if department not in DeptEnum.__members__:
        raise HTTPException(400, f"Unknown department: {department}")
    dept = DeptEnum[department]

    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)")

    import tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        parsed = parse_excel(tmp_path)
    except Exception as e:
        raise HTTPException(422, f"Parse error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not parsed:
        raise HTTPException(422, "No data rows found. Check file format.")

    # Один kod может повторяться в файле (разные месяцы) с одним и тем же
    # именем — оставляем последнее непустое имя на код.
    name_by_kod: dict[str, str] = {}
    for r in parsed:
        if r["kod"] and r["name"]:
            name_by_kod[r["kod"]] = r["name"]

    updated = 0
    already_ok = 0
    not_found = []
    for kod, correct_name in name_by_kod.items():
        result = await db.execute(
            select(KaspiRow).where(
                KaspiRow.department == dept,
                KaspiRow.kod == kod,
            )
        )
        rows = result.scalars().all()
        if not rows:
            not_found.append(kod)
            continue
        for row in rows:
            if not row.name or row.name.strip() == "" or row.name.strip() == kod.strip():
                row.name = correct_name
                updated += 1
            else:
                already_ok += 1

    await db.commit()

    return {
        "department": department,
        "kods_in_file": len(name_by_kod),
        "rows_updated": updated,
        "rows_already_had_a_name": already_ok,
        "kods_not_found_in_db": len(not_found),
        "kods_not_found_sample": not_found[:20],
    }


@router.get("/", summary="List uploads")
async def list_uploads(
    department: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    q = select(Upload).order_by(Upload.created_at.desc())
    if department and department in DeptEnum.__members__:
        q = q.where(Upload.department == DeptEnum[department])
    result = await db.execute(q)
    uploads = result.scalars().all()
    return [
        {
            "id": u.id,
            "filename": u.filename,
            "department": u.department.value,
            "row_count": u.row_count,
            "months": u.months,
            "subtypes": u.subtypes,
            "created_at": u.created_at.isoformat(),
        }
        for u in uploads
    ]


@router.get("/unclassified", summary="Rows with no Тип that need a manual decision")
async def list_unclassified(
    department: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Products currently sitting in the DB with an empty Тип (tip_classifier.py
    couldn't resolve them with hard evidence). Grouped by kod, since the same
    product can repeat across months — one decision fixes every row for it,
    this month and every month going forward (the classifier will exact-kod
    match it next time).
    """
    if department not in DeptEnum.__members__:
        raise HTTPException(400, f"Unknown department: {department}")
    dept = DeptEnum[department]

    from sqlalchemy import func
    q = (
        select(
            KaspiRow.kod,
            KaspiRow.name,
            KaspiRow.brand,
            func.sum(KaspiRow.revenue).label("revenue"),
            func.count(KaspiRow.id).label("rows"),
        )
        .where(
            KaspiRow.department == dept,
            (KaspiRow.tip.is_(None)) | (KaspiRow.tip == ""),
        )
        .group_by(KaspiRow.kod, KaspiRow.name, KaspiRow.brand)
        .order_by(func.sum(KaspiRow.revenue).desc())
    )
    rows = (await db.execute(q)).all()
    return {
        "department": department,
        "count": len(rows),
        "total_revenue": sum(r.revenue or 0 for r in rows),
        "items": [
            {"kod": r.kod, "name": r.name, "brand": r.brand, "revenue": r.revenue or 0, "row_count": r.rows}
            for r in rows
        ],
    }


@router.patch("/tip", summary="Set Тип for one product (applies to every row with this kod)")
async def set_tip(
    department: str,
    kod: str,
    tip: str,
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_admin),
):
    """
    One-time decision that becomes permanent: sets Тип on every existing row
    for this kod+department, AND — because tip_classifier.py's exact_kod tier
    always trusts the DB's own history first — every future month's upload of
    the same product resolves automatically from here on. No file to
    re-upload, no need to remember the decision next time.
    """
    if department not in DeptEnum.__members__:
        raise HTTPException(400, f"Unknown department: {department}")
    dept = DeptEnum[department]
    if not tip.strip():
        raise HTTPException(400, "tip cannot be empty")

    from sqlalchemy import update
    result = await db.execute(
        update(KaspiRow)
        .where(KaspiRow.department == dept, KaspiRow.kod == kod)
        .values(tip=tip.strip())
    )
    await db.commit()
    return {"kod": kod, "tip": tip.strip(), "rows_updated": result.rowcount}


@router.delete("/{upload_id}", summary="Delete upload and its rows")
async def delete_upload(upload_id: int, db: AsyncSession = Depends(get_db), _: None = Depends(require_admin)):
    upload = await db.get(Upload, upload_id)
    if not upload:
        raise HTTPException(404, "Upload not found")

    # Delete the file from disk (uploaded files are named *_{original_filename})
    if upload.filename:
        for f in glob.glob(os.path.join(UPLOAD_DIR, f"*_{upload.filename}")):
            try:
                os.remove(f)
            except OSError:
                pass

    await db.delete(upload)
    await db.commit()
    return {"deleted": upload_id}
