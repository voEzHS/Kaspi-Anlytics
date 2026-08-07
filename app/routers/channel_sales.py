"""«Продажи всех каналов» (torgstore.zymyran.com, транзакционная выгрузка
CRM-аналитики — «Экспорт → Аналитика → Товары», custom-отчёт) — загрузка
снимка продаж по SKU/дате/каналу/категории, используется движком
calc_procurement_v2 (см. app/analytics/engine.py и Zakup_V2_Design_2026-08-05.md
в корне репозитория — дизайн-документ с обоснованием каждого решения ниже).

Формат подтверждён на реальном файле 05.08.2026: колонки — SKU, Название
товара, Количество, Сумма, Номер накладной, Номер заявки Kaspi, Имя клиента,
Номер телефона, Дата заявки, Дата отпуска, Месяц отпуска, Канал продаж,
Город канала продаж, Город накладной, Подгруппа товара, Категория товара.
Клиентские ПДн (имя, телефон) сознательно НЕ читаются и не сохраняются —
для процурмента не нужны.

08.08 — «Город канала продаж» теперь читается (поле city). Проверено на
реальном файле: ровно 3 значения без пропусков — Алматы/Астана/Шымкент,
100% совпадает со складами StockRow (wh_pervomay/wh_astana/wh_shymkent) —
это город склада, который обслужил продажу, а не город доставки клиенту.
Именно поэтому он пригоден для разбивки закупа/перемещения по городам.
«Город накладной» (город доставки клиента, 285 разных значений вплоть до
посёлков) — сознательно НЕ читаем: это другое измерение (география спроса
конечного клиента по всему Казахстану), не совпадает со складами и не
нужен для решений по закупу/перемещению между 3 складами.
"""
import os
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.models import ChannelSalesRow, ChannelSalesUpload
from app.routers.uploads import require_admin, MAX_UPLOAD_BYTES

router = APIRouter(prefix="/api/v1/channel-sales", tags=["channel-sales"])

_COL_SKU = "sku"
_COL_NAME = "название"
_COL_QTY = "количество"
_COL_REVENUE = "сумма"
_COL_DATE = "дата отпуска"
_COL_CHANNEL = "канал продаж"
_COL_CATEGORY = "категория товара"
_COL_SUBGROUP = "подгруппа товара"
_COL_CITY = "город канала продаж"

# Нормализация написания города к 3 каноничным значениям, совпадающим с
# StockRow.wh_pervomay/wh_astana/wh_shymkent. В файле встречается ровно эти
# 3 варианта без пропусков, но нормализуем на всякий случай (регистр/пробелы).
_CITY_CANON = {"алматы": "Алматы", "астана": "Астана", "шымкент": "Шымкент"}


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def _parse_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_channel_sales_excel(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return []

    header_idx = 0
    for i, row in enumerate(rows_raw[:5]):
        if any(_norm(c) == _COL_SKU for c in row):
            header_idx = i
            break
    headers = rows_raw[header_idx]

    col = {}
    for idx, h in enumerate(headers):
        n = _norm(h)
        if not n:
            continue
        if n == _COL_SKU:
            col["sku"] = idx
        elif _COL_NAME in n and "товар" in n:
            col["name"] = idx
        elif n == _COL_QTY:
            col["qty"] = idx
        elif n == _COL_REVENUE:
            col["revenue"] = idx
        elif n == _COL_DATE:
            col["date"] = idx
        elif n == _COL_CHANNEL:
            col["channel"] = idx
        elif n == _COL_CATEGORY:
            col["category"] = idx
        elif n == _COL_SUBGROUP:
            col["subgroup"] = idx
        elif n == _COL_CITY:
            col["city"] = idx

    if "sku" not in col:
        raise ValueError("Колонка SKU не найдена в файле")

    def get(row, key):
        i = col.get(key)
        return row[i] if i is not None and i < len(row) else None

    result = []
    for row in rows_raw[header_idx + 1:]:
        if not row or col["sku"] >= len(row):
            continue
        sku_raw = get(row, "sku")
        if sku_raw is None or str(sku_raw).strip() == "":
            continue
        sku = str(sku_raw).strip().upper()

        dt = _parse_dt(get(row, "date"))
        city_raw = _norm(get(row, "city"))
        result.append({
            "sku": sku,
            "name": str(get(row, "name") or "").strip(),
            "qty": _num(get(row, "qty")),
            "revenue": _num(get(row, "revenue")),
            "sale_date": dt,
            "channel": str(get(row, "channel") or "").strip(),
            "category": str(get(row, "category") or "").strip(),
            "subgroup": str(get(row, "subgroup") or "").strip(),
            "city": _CITY_CANON.get(city_raw),  # None если незнакомое написание — не угадываем
        })
    return result


@router.post("/", summary="Загрузить снимок продаж по всем каналам (заменяет предыдущий целиком)")
async def upload_channel_sales(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_admin),
):
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)")

    import tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        parsed = parse_channel_sales_excel(tmp_path)
    except Exception as e:
        raise HTTPException(422, f"Ошибка разбора файла: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not parsed:
        raise HTTPException(422, "В файле не найдено строк с данными")

    # Не накопительно — новый экспорт содержит полную историю заново (см.
    # docstring модуля и ChannelSalesUpload в models.py).
    old_uploads = (await db.execute(select(ChannelSalesUpload))).scalars().all()
    for u in old_uploads:
        await db.delete(u)
    await db.flush()

    upload = ChannelSalesUpload(filename=file.filename, row_count=len(parsed))
    db.add(upload)
    await db.flush()

    CHUNK = 1000
    for i in range(0, len(parsed), CHUNK):
        chunk = parsed[i: i + CHUNK]
        db.add_all([ChannelSalesRow(upload_id=upload.id, **row) for row in chunk])

    await db.commit()
    await db.refresh(upload)

    return {
        "id": upload.id,
        "filename": upload.filename,
        "row_count": upload.row_count,
        "created_at": upload.created_at.isoformat(),
        "replaced_previous": len(old_uploads) > 0,
    }


@router.get("/", summary="Текущий снимок продаж по каналам (метаданные)")
async def get_channel_sales_status(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ChannelSalesUpload).order_by(ChannelSalesUpload.created_at.desc()))
    upload = result.scalars().first()
    if not upload:
        return {"loaded": False}
    return {
        "loaded": True,
        "id": upload.id,
        "filename": upload.filename,
        "row_count": upload.row_count,
        "created_at": upload.created_at.isoformat(),
    }


@router.delete("/", summary="Удалить текущий снимок продаж по каналам")
async def delete_channel_sales(db: AsyncSession = Depends(get_db), _: None = Depends(require_admin)):
    old_uploads = (await db.execute(select(ChannelSalesUpload))).scalars().all()
    for u in old_uploads:
        await db.delete(u)
    await db.commit()
    return {"deleted": len(old_uploads)}
